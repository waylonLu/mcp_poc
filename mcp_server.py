from fastmcp import FastMCP
from typing import Dict, Any, List
from db.database import db
from clients.api_client import api_client
from schemas.financial_models import FinancialProduct, UserInvestment
import sqlite3
import asyncio
import json
import os
import openpyxl
import httpx
from dotenv import load_dotenv

load_dotenv()

# Create FastMCP server instances
mcp = FastMCP(name="api-integration-server")
mcp_banking = FastMCP(name="banking-server")

# @mcp.tool(name="get_cherrypicks_info", 
#           description="cherrypicks(创奇思科技有限公司) is the name of a company. This api returns the company's policies"
#           )
# async def get_cherrypicks_info(query: str) -> Dict[str, Any]:
#     """
#     Get cherrypicks info
#     """
#     try:
#         params ={
#             "inputs": {},
#             "query": query,
#             # "response_mode": "streaming",
#             "response_mode": "blocking",
#             "conversation_id": "",
#             "user": "AI_Agent",
#             "files": []
#         }
                    
#         result = await api_client.make_request("get_cherrypicks_info", params)
#         return result
#     except Exception as e:
#         return {
#             "status": "error",
#             "message": f"Get cherrypicks info failed: {str(e)}"
#         }

# 转账功能
@mcp_banking.tool(
    name="transfer_money",
    description="Transfer money from one account to another. Args: from_account_id (source account), to_account_id (destination account), amount (transfer amount), description (optional transfer note). Returns a result message."
)
def transfer_money(from_account_id: str, to_account_id: str, amount: float, description: str = None) -> str:
    """Transfer function
    
    Args:
        from_account_id: Source account ID
        to_account_id: Destination account ID
        amount: Transfer amount
        description: Transfer description (optional)
    """
    # Check if accounts exist
    from_account = db.get_account(from_account_id)
    to_account = db.get_account(to_account_id)
    
    if not from_account:
        return f"Error: Source account {from_account_id} does not exist"
    
    if not to_account:
        return f"Error: Destination account {to_account_id} does not exist"
    
    # Check if amount is valid
    if amount <= 0:
        return f"Error: Transfer amount must be greater than 0"
    
    # Check if balance is sufficient
    if from_account.balance < amount:
        return f"Error: Insufficient balance. Current balance: {from_account.balance}"
    
    # Execute transfer
    db.update_balance(from_account_id, -amount)
    db.update_balance(to_account_id, amount)
    tran = db.add_transaction(from_account_id, to_account_id, amount, description)
    
    return f"Transfer successful, transaction ID: {tran.id}! From account: {format_card_number(from_account.card_number)} to account: {format_card_number(to_account.card_number)} amount: {amount} RMB"

# 查询余额功能
@mcp_banking.tool(
    name="check_balance",
    description="Check the balance of a specific account. Args: account_id (account to check). Returns the current balance or error message."
)
def check_balance(account_id: str) -> str:
    """Check account balance
    
    Args:
        account_id: Account ID
    """
    account = db.get_account(account_id)
    
    if not account:
        return f"Error: Account {account_id} does not exist"
    
    return f"Account name: ({account.name})\nCard number: {format_card_number(account.card_number)} \nCurrent balance: {account.balance} RMB"

# 查询账户信息
@mcp_banking.tool(
    name="get_account_info",
    description="Get information of a specific account. Args: account_name (account to query). Returns account id, name or error message."
)
def get_account_info(account_name: str) -> str:
    """Query account information
    
    Args:
        account_name: Account name
    """
    account = db.get_account_by_name(account_name)
    
    if not account:
        return f"Error: Account {account_name} does not exist"
    
    return f"Account ID: {account.id}\nAccount name: {account.name}\n Card number: {format_card_number(account.card_number)}"

# 查询交易历史功能
@mcp_banking.tool(
    name="get_transaction_history",
    description="Get recent transaction history for an account. Args: account_id (account to query), limit (number of records, default 10). Returns formatted transaction list or error message."
)
def get_transaction_history(account_id: str, limit: int = 10) -> str:
    """Query account transaction history
    
    Args:
        account_id: Account ID
        limit: Number of records to return (default 10)
    """
    account = db.get_account(account_id)
    
    if not account:
        return f"Error: Account {account_id} does not exist"
    
    transactions = db.get_account_transactions(account_id, limit)
    
    if not transactions:
        return f"Account {account_id} has no transaction records"
    
    result = [f"Account {account_id} ({account.name}) transaction history (latest {len(transactions)} records):"]
    for i, t in enumerate(transactions, 1):
        if t.from_account == account_id:
            direction = "Outgoing"
            other_account = t.to_account
        else:
            direction = "Incoming"
            other_account = t.from_account
        
        result.append(
            f"{i}. {t.timestamp.strftime('%Y-%m-%d %H:%M:%S')} "
            f"{direction} {other_account} {t.amount} RMB "
            f"{t.description or ''}"
        )
    
    return "\n".join(result)

# 列出所有账户功能
@mcp_banking.tool(
    name="list_accounts",
    description="List all accounts in the system. Returns a formatted list of account IDs, names, and balances."
)
def list_accounts() -> str:
    """List all accounts"""
    try:
        with db.get_connection() as conn:
            cursor = conn.execute("SELECT id, name, balance FROM accounts ORDER BY id")
            accounts = cursor.fetchall()
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return []
    
    if not accounts:
        return "No accounts available"
    
    result = ["All accounts:"]
    for account in accounts:
        result.append(f"- {account[0]}: {account[1]} (Balance: {account[2]} RMB)")
    
    return "\n".join(result)

def format_card_number(card_number: str) -> str:
    """
    Format bank card number as 6217...1234, showing only the first 4 and last 4 digits.
    Example: '6222021234567890123' -> '6222...0123'
    """
    if not card_number or len(card_number) < 8:
        return card_number or ""
    return f"{card_number[:4]}...{card_number[-4:]}"


# 填写报销表格
_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template", "报销表格_202501.xlsx")
_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "template", "output")

LOCAL_CATEGORIES = ["餐食", "交通", "娱乐", "医疗费", "通讯费", "电子相关", "其它费用"]
OVERSEAS_CATEGORIES = ["旅费报销(机票)", "旅费报销(住宿费)", "旅费报销(车费)", "旅费报销(餐费)", "其它费用", "通讯费"]

@mcp.tool(
    name="fill_expense_report",
    description="""当用户需要填写报销表、生成报销单、提交费用报销、制作报销 Excel 文件时，调用此工具。
Use this tool when the user wants to fill in an expense report, generate a reimbursement form, or submit expense claims.

This tool fills the Excel expense report template (报销表格_202501.xlsx) and uploads it to the file server.

Parameters:
- name: Employee name (姓名)，必填
- period: Expense period（报销时期），e.g. '2025年1月'，必填
- sheet_type: '本地' for local expenses（默认）or '海外' for overseas expenses
- items: JSON array string of expense items（费用明细），必填。
  本地 format: [{"date": "2025-01-15", "details": "客户餐饮", "category": "餐食", "amount": 150.0}, ...]
  Valid 本地 categories: 餐食 / 交通 / 娱乐 / 医疗费 / 通讯费 / 电子相关 / 其它费用
  海外 format: [{"date": "2025-01-15", "details": "Flight ticket", "category": "旅费报销(机票)", "amount": 500.0}, ...]
  Valid 海外 categories: 旅费报销(机票) / 旅费报销(住宿费) / 旅费报销(车费) / 旅费报销(餐费) / 其它费用 / 通讯费
  (海外 amounts are in original currency; RMB conversion is auto-calculated via exchange rate)
- project_name: Project name（项目名称，海外 only）
- original_currency: Currency code（币种），e.g. 'USD' or 'HKD'（海外 only）
- exchange_rate: Exchange rate to RMB（汇率，海外 only，e.g. 7.2 for USD）
- output_filename: Output filename without extension（输出文件名，可选，不填则自动生成）

Returns a summary with the upload URL of the generated Excel file."""
)
def fill_expense_report(
    name: str,
    period: str,
    items: str,
    sheet_type: str = "本地",
    project_name: str = "",
    original_currency: str = "USD",
    exchange_rate: float = 7.2,
    output_filename: str = ""
) -> str:
    missing = [f for f, v in [("name", name), ("period", period), ("items", items)] if not v or not str(v).strip()]
    if missing:
        return f"Error: 以下必填字段缺失，请补充后重试：{', '.join(missing)}"

    if sheet_type not in ("本地", "海外"):
        return "Error: sheet_type must be '本地' or '海外'"

    try:
        expense_items = json.loads(items)
    except json.JSONDecodeError as e:
        return f"Error: Invalid items JSON — {str(e)}"

    if not isinstance(expense_items, list):
        return "Error: items must be a JSON array"

    try:
        wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    except Exception as e:
        return f"Error: Cannot load template file — {str(e)}"

    if sheet_type == "本地":
        ws = wb["报销表格_本地"]
        ws["B5"] = name
        ws["B6"] = period
        start_row = 9
        max_rows = 30  # rows 9–38
        for i, item in enumerate(expense_items[:max_rows]):
            row = start_row + i
            ws[f"A{row}"] = item.get("date", "")
            ws[f"B{row}"] = item.get("details", "")
            ws[f"C{row}"] = item.get("category", "")
            ws[f"D{row}"] = item.get("amount", 0)
    else:  # 海外
        ws = wb["报销表格_海外"]
        ws["B4"] = name
        ws["B5"] = period
        ws["B6"] = project_name
        ws["L6"] = original_currency
        ws["L7"] = exchange_rate
        start_row = 9
        max_rows = 34  # rows 9–42 (row 43 starts local-currency section)
        for i, item in enumerate(expense_items[:max_rows]):
            row = start_row + i
            ws[f"A{row}"] = item.get("date", "")
            ws[f"B{row}"] = item.get("details", "")
            ws[f"C{row}"] = item.get("category", "")
            ws[f"D{row}"] = item.get("amount", 0)
            # Column E (RMB amount) is auto-calculated by the existing formula =ROUND(D*$L$7,2)

    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    if not output_filename:
        safe_period = period.replace("/", "-").replace(" ", "_")
        output_filename = f"报销表格_{name}_{safe_period}_{sheet_type}"
    output_path = os.path.join(_OUTPUT_DIR, f"{output_filename}.xlsx")

    try:
        wb.save(output_path)
    except Exception as e:
        return f"Error: Cannot save output file — {str(e)}"

    # 上传到文件服务
    upload_info = ""
    upload_url = os.getenv("UPLOAD_API_URL")
    if upload_url:
        try:
            with open(output_path, "rb") as f:
                with httpx.Client(timeout=30) as client:
                    upload_resp = client.post(
                        upload_url,
                        files={"file": (f"{output_filename}.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                        data={"filePath": "agent_source", "rename": "false"},
                    )
            resp_json = upload_resp.json()
            if resp_json.get("code") == "000000" and resp_json.get("data"):
                file_url = resp_json["data"][0].get("url", "")
                os.remove(output_path)
                print(f"[上传成功] {output_filename}.xlsx -> {file_url}")
                upload_info = f"\n文件已上传至服务器\n下载链接: {file_url}"
            else:
                msg = resp_json.get('msg', '未知错误')
                print(f"[上传失败] {output_filename}.xlsx -> 服务器返回: {msg}")
                upload_info = f"\n文件上传失败: {msg}"
        except Exception as e:
            print(f"[上传失败] {output_filename}.xlsx -> 异常: {str(e)}")
            upload_info = f"\n文件上传失败: {str(e)}"

    filled = min(len(expense_items), max_rows)
    saved_info = "" if upload_info.startswith("\n文件已上传") else f"\n文件已保存至: {output_path}"
    return (
        f"报销表格已填写完成！\n"
        f"{upload_info}"
    )

async def run_mcp():
    await mcp.run_async(transport="sse", host="0.0.0.0")

async def run_mcp_banking():
    await mcp_banking.run_async(transport="sse", port=8001, host="0.0.0.0")

# 理财产品查询功能
@mcp_banking.tool(
    name="list_financial_products",
    description="List all available financial products. Returns a formatted list of products with details."
)
def list_financial_products() -> str:
    """List all available financial products"""
    products = db.get_financial_products()
    
    if not products:
        return "No financial products available"
    
    result = ["Available financial products:"]
    for i, product in enumerate(products, 1):
        result.append(
            f"{i}. {product.name} (ID: {product.id})\n"
            f"   Description: {product.description}\n"
            f"   Minimum investment: {product.min_investment} RMB\n"
            f"   Expected return rate: {product.expected_return_rate * 100:.1f}%\n"
            f"   Risk level: {product.risk_level}\n"
            f"   Duration: {product.duration_days} days"
        )
    
    return "\n".join(result)

# 购买理财产品功能
@mcp_banking.tool(
    name="purchase_financial_product",
    description="Purchase a financial product. Args: account_id (account to purchase from), product_id (product ID to purchase), amount (investment amount). Returns purchase result message."
)
def purchase_financial_product(account_id: str, product_id: str, amount: float) -> str:
    """Purchase financial product
    
    Args:
        account_id: Account ID to purchase from
        product_id: Financial product ID
        amount: Investment amount
    """
    investment = db.purchase_financial_product(account_id, product_id, amount)
    
    if not investment:
        return "Purchase failed. Please check: 1) Account exists, 2) Product exists, 3) Amount meets minimum requirement, 4) Sufficient balance"
    
    return f"Purchase successful! Investment ID: {investment.id}\n" \
           f"Amount: {amount} RMB\n" \
           f"Expected maturity date: {investment.expected_maturity_date.strftime('%Y-%m-%d')}\n" \
           f"Status: {investment.status}"

# 查询用户投资记录功能
@mcp_banking.tool(
    name="get_user_investments",
    description="Get investment records for a specific account. Args: account_id (account to query). Returns formatted investment list or error message."
)
def get_user_investments(account_id: str) -> str:
    """Get user investment records
    
    Args:
        account_id: Account ID
    """
    account = db.get_account(account_id)
    if not account:
        return f"Error: Account {account_id} does not exist"
    
    investments = db.get_user_investments(account_id)
    
    if not investments:
        return f"Account {account_id} has no investment records"
    
    result = [f"Investment records for account {account_id} ({account.name}):"]
    for i, investment in enumerate(investments, 1):
        product = db.get_financial_product(investment.product_id)
        product_name = product.name if product else investment.product_id
        
        result.append(
            f"   Amount: {investment.investment_amount} RMB\n"
            f"   Investment date: {investment.investment_date.strftime('%Y-%m-%d')}\n"
            f"   Expected maturity date: {investment.expected_maturity_date.strftime('%Y-%m-%d')}\n"
            f"   Status: {investment.status}"
        )
    
    return "\n".join(result)

async def main():
    # Run both services concurrently
    await asyncio.gather(run_mcp(), run_mcp_banking())

if __name__ == "__main__":
    asyncio.run(main())
